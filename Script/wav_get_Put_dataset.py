'''
Lo script prende in input il file csv creato dallo script wavExtractor e va a confrontare i nomi delle istanze con
quelli del foglio excel smistamenti e a sencoda della colonna oin cui li trova in quest'ultimo va ad assegnare
 ad un'istanza l'etichetta get put e crea il file arff risultante

 NB: nel file excell smistamenti ci sono alcune stringhe che hanno * alla fine vanno eleiminate altirmenti weka non riconosce il file

'''
import pandas as pd
import openpyxl

'''
conda install pandas
conda install xlrd 
conda install openpyxl
'''
# apro file smistamento
load_smisted_file = openpyxl.load_workbook("Smistamento_Dataset_Acid.xlsx")
smisted_file = load_smisted_file.active

'''
 apriamo il file csv generato dall'estrazione dei segmenti, il file con le bags 
 Deve avere nella prima colonna i nomi (bag id)
 nella seconda colonna le features (bag) tra " "
 nell'ultima il tipo di risorsa (classe)
 
 Non deve avere la colonna con gli indici delle colonne ma la prima riga dev'essere vuota
'''
dataset_csv = pd.read_csv('results\\data.csv')

'''
Usiamo le funzioni definite nello script wavDatasetLib per creare il file .arff
Usimo il replace per non creare conflitti in weka nell'attributo @class
'''
import wavDatasetLib

put_type = str(smisted_file.cell(row=1, column=3).value).replace(" ", "_")  # put
get_type = str(smisted_file.cell(row=1, column=4).value).replace(" ", "_")  # get
wavDatasetLib.createArff("dataGetPut", "virus", wavDatasetLib.getHeaderAttributes(), "real", [get_type, put_type])
fileArff = open("results\\dataGetPut.arff", 'a', newline='')
fileArff.close()
'''
@valueID: stringe del nome dell'ima applicazione
@:return una strigna che identifica l'applicazione come PUT o GET 

La fuzione prende in  input una stringa e ricerca nel file .xlsx se è presente in elenco nella 3 o 4 colonna 
rispettivamente sapremo che è un app di PUT o di GET che sappiamo trovarsi alla prima riga del file.
'''


def getTypeOfApp(valueID):
    for i in range(1, smisted_file.max_row + 1):  # itero sulle righe

        # controllo se il nome in esame conincide con quello passato
        if valueID == smisted_file.cell(row=i, column=3).value:
            print("set: ", smisted_file.cell(row=i, column=1).value, " ", valueID, ": ",
                  smisted_file.cell(row=i, column=3).value, ": ",
                  smisted_file.cell(row=1, column=3).value)
            return str(smisted_file.cell(row=1, column=3).value)  # put

        elif valueID == smisted_file.cell(row=i, column=4).value:
            print("set: ", smisted_file.cell(row=i, column=1).value, " ", valueID, ": ",
                  smisted_file.cell(row=i, column=4).value, ": ",
                  smisted_file.cell(row=1, column=4).value)
            return str(smisted_file.cell(row=1, column=4).value)  # get
        else:
            continue

#TODO: inserire i dati in results

for index, row in dataset_csv.iterrows():  # iteriamo sulle righe
    typeWav = row[2]  # ritorna ultima colonna con il tipo della risorsa
    nameWav = row[0]
    if typeWav == "trusted":
        continue
    else:
        # estraiamo il tipo di risorsa passando alla funzione il nome dell'istanza senza il .wav
        type = str(getTypeOfApp(str(row[0]).split(".")[0])).replace(" ", "_")

        # creiamo la stringa da inserire nel file .arff
        # row[1] è la bag
        stringa = f'{nameWav}{","}\"{row[1]}\"{","}{type}\n'

        # print(int, " *******", len(stringa.split(",")))

        fileArff = open("results\\dataGetPut.arff", 'a', newline='')
        fileArff.writelines(stringa)
        fileArff.close()

'''
Non posso aprire il file .arff perche MIL 
'''
# data, meta = arff.loadarff(open('data.arff'))
# df = pd.DataFrame(features_data_arff[0])
# df.head()
